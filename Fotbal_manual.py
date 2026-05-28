import openpyxl
from datetime import datetime

def genereaza_raport_excel_manual():
    print("🤖 MULTI-SCANNER - INTRODUCERE MANUALĂ COTE & EXPORT EXCEL")
    print("=============================================================")
    print("Introduceți datele meciurilor. Pentru a opri și genera fișierul Excel, tastați 'exit' la numele meciului.\n")
    
    meciuri_introduse = []
    
    while True:
        meci_nume = input("⚽ Meci (ex: Malmo FF vs Halmstads sau 'exit'): ").strip()
        if meci_nume.lower() == 'exit':
            break
            
        liga = input("🏆 Competiție / Ligă (ex: Allsvenskan - Sweden): ").strip()
        ora_meci = input("⏰ Dată / Oră (ex: 20:30 / lăsați gol pentru ora curentă): ").strip()
        if not ora_meci:
            ora_meci = datetime.now().strftime("%H:%M")
            
        try:
            c1 = float(input("📉 Cotă 1 (Gazde): "))
            cX = float(input("📉 Cotă X (Egal - puneți 0 dacă este Baschet): "))
            c2 = float(input("📉 Cotă 2 (Oaspeți): "))
        except ValueError:
            print("⚠️ Eroare: Cotele trebuie să fie numere cu punct (ex: 1.65). Reluați acest meci.\n")
            continue

        if " vs " in meci_nume:
            gazde, oaspeti = meci_nume.split(" vs ", 1)
        else:
            gazde, oaspeti = meci_nume, "Oaspeți"

        meciuri_introduse.append({
            'ora': ora_meci,
            'liga': liga,
            'meci_complet': meci_nume,
            'g': gazde,
            'o': oaspeti,
            'c1': c1, 'cX': cX, 'c2': c2
        })
        print("🟩 Meci salvat temporar! Treceți la următorul.\n")

    if not meciuri_introduse:
        print("⚠️ Nu ați introdus niciun meci. Fișierul Excel nu a fost creat.")
        return

    print("\n📊 Generăm registrul Excel și aplicăm algoritmii de predicție...")
    wb = openpyxl.Workbook()
    
    ws_general = wb.active
    ws_general.title = "Panou General Rezultate"
    
    antet_general = [
        "Dată / Oră", "Competiție", "Meci", "Cota 1", "Cota X", "Cota 2", 
        "Verdict Risc", "🛡️ Varianta Safe", "⚡ Varianta Medium", "🔥 Varianta Risk"
    ]
    ws_general.append(antet_general)

    ws_bilet = wb.create_sheet(title="Recomandări Bilet VIP")
    antet_bilet = ["Dată / Oră", "Competiție", "Meci", "Tip Recomandare", "Opțiune Sugerată", "Cota 1"]
    ws_bilet.append(antet_bilet)

    contor_verzi = 0
    
    for m in meciuri_introduse:
        c1, cX, c2 = m['c1'], m['cX'], m['c2']
        este_baschet = (cX == 0)
        
        # --- ALGORITMUL MATEMATIC MODIFICAT ---
        if c1 <= 1.65:
            verdict = "🟩 FAVORIT CLAR"
            sugestie_safe = "Câștigător Meci" if este_baschet else "Șansă dublă 1X (Asigurat)"
            # 🔥 AICI AM ADĂUGAT OPȚIUNEA PENTRU PRIMA REPRIZĂ LA FAVORIȚI
            sugestie_medium = "Victorie la Handicap" if este_baschet else "⚽ Gol în Prima Repriză (Peste 0.5)"
            sugestie_risk = "Peste prag puncte" if este_baschet else "1 & Peste 2.5 Goluri"
            
            # Adăugăm în biletul de siguranță (Filtrul Verde)
            ws_bilet.append([m['ora'], m['liga'], m['meci_complet'], "🛡️ Bază Safe", sugestie_safe, f"{c1:.2f}"])
            ws_bilet.append([m['ora'], m['liga'], m['meci_complet'], "⚡ Cotă Medie", sugestie_medium, f"{c1:.2f}"])
            contor_verzi += 1
            
        elif 1.65 < c1 <= 2.30:
            verdict = "🟨 MECI ECHILIBRAT"
            sugestie_safe = "Total puncte mediu" if este_baschet else "Interval goluri 1-4"
            # 🔥 ȘI AICI FUNCȚIONEAZĂ EXCELENT „AMBELE MARCHEAZĂ” SAU „GOL ÎN PRIMA REPRIZĂ”
            sugestie_medium = "Victorie în prelungiri" if este_baschet else "⚽ Gol în 1R (Peste 0.5) SAU GG"
            sugestie_risk = "Scor strâns" if este_baschet else "X final (Egalitate)"
            
        else:
            verdict = "🟥 OUTSIDER / HAZARD"
            sugestie_safe = "Handicap oaspeți" if este_baschet else "Peste 1.5 Goluri în meci"
            sugestie_medium = "Victorie oaspeți" if este_baschet else f"Șansă dublă X2 ({m['o']})"
            sugestie_risk = "Surpriză mare" if este_baschet else f"2 Solist ({m['o']})"

        cota_X_afisare = "N/A (Baschet)" if este_baschet else f"{cX:.2f}"

        ws_general.append([
            m['ora'], m['liga'], m['meci_complet'], 
            f"{c1:.2f}", cota_X_afisare, f"{c2:.2f}",
            verdict, sugestie_safe, sugestie_medium, sugestie_risk
        ])

    # Auto-ajustare dimensiuni coloane
    for ws in [ws_general, ws_bilet]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Salvare fișier local final
    data_azi = datetime.now().strftime("%d_%m_%Y")
    nume_excel = f"Predictii_Excel_Manual_{data_azi}.xlsx"
    wb.save(nume_excel)
    
    print("\n=============================================================")
    print(f"✅ FIȘIERUL EXCEL A FOST GENERAT: {nume_excel}")
    print(f"🔹 Meciuri totale introduse de tine: {len(meciuri_introduse)}")
    print(f"🔹 Meciuri prinse în Filtrul Verde (Sigure): {contor_verzi}")
    print("=============================================================")

if __name__ == "__main__":
    genereaza_raport_excel_manual()